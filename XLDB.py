import os
from shadok.style import OK,FAIL
from functools import wraps


def try_import(install_tries:int=1):
    """Tries to import openpyxl, if fail, tries to install it by running os.system("pip install openpyxl")

    Args:
        install_tries (int, optional): How many times do we call 'pip install ...' before resigning. Defaults to 1.
    """
    try:
        print("Trying to import openpyxl...",end=" ")
        import openpyxl
        OK()
    except:
        FAIL()
        print()
        if (install_tries==0):
            return
        print("Installing openpyxl...")
        os.system("pip install openpyxl")
        try_import(install_tries-1)
        
try_import()

from openpyxl import Workbook as Wb, worksheet as Ws, load_workbook

class TableNameException(Exception):
    """table nam can't contain any spaces or be equal to "Sheet"
    """
    def __init__(self) -> None:
        super().__init__("Invalid table name, please choose somthing else !")

class IncompleteClassDefinition(Exception):
    def __init__(self,missing_method:str) -> None:
        super().__init__("Missing method : "+missing_method)

class InvalidID(Exception):
    def __init__(self):
        super().__init__("Inexistant ID in database")

def saveDB(func):
    """wrapper that allows you to run the function, save the database and then return the result of the function

    Args:
        func: wrapped function

    Returns:
        function: wrapper
    """
    @wraps(func)
    def wrapper(*args,**kwargs):
        out = func(*args,**kwargs)
        Database.save()
        return out
    return wrapper
    
    
class Database:
    
    name:str = "Database"
    wb:Wb = None
    
    @staticmethod
    def table_exists(table_name:str)->bool:
        return table_name in Database.wb.sheetnames
    
    @staticmethod
    @saveDB
    def add_table(table_name:str)->None:
        if table_name=="Sheet" or " " in table_name:
            raise TableNameException()
        
        if not Database.table_exists(table_name):
            Database.wb.create_sheet(table_name)
    
    @staticmethod
    def get_table(table_name)->Ws:
        return Database.wb[table_name]
   
    @staticmethod 
    @saveDB   
    def drop_table(table_name:str):
        """erase all the content of the table but without removing the table
        """
        Database.delete_table(table_name)
        Database.wb.create_sheet(table_name)
    
    @staticmethod
    @saveDB
    def delete_table(table_name:str):
        """delete tabme from database"""
        Database.wb.remove(Database.get_table(table_name))
        
    @staticmethod
    def save()->None:
        """save database to location"""
        Database.wb.save(Database.filename())
    
    @staticmethod
    def start(name:str="XLdatabase")->None:
        """create database if doesn't exists and updates Database.wb"""
        Database.name = name
        print(f"Starting Database {Database.filename()}")
        if not os.path.exists(Database.filename()):
            Database.wb = Wb()
            Database.save()
        else:
            Database.wb = load_workbook(Database.filename())
    
    @staticmethod
    @saveDB
    def reset()->None:
        """recreate an empty database"""
        Database.wb = Wb()
    
    @staticmethod
    def filename()->str:
        return Database.name+".xlsx"


class Filter:
    
    def __init__(self,*args,**kwargs):
        """Args:
            function (function, optional) : function(cls_obj) -> bool
            There can be multiple functions (AND operator)
            
           Kwargs:
            properties comparaison (ex : age = 31) (if age=None, age is not verified)
        """
        self.functions = args
        self.property_checker = kwargs
    
    def __call__(self, cls_obj) -> bool:
        for function in self.functions:
            if not function(cls_obj):
                return False
        
        for key,value in self.property_checker.items():
            if cls_obj.__getattribute__(key) != value and value!=None:
                return False
        
        return True
    
    
class Model:
    """Required : get_property_names() (@staticmethod)
    """
    
    def __init__(self,id:int=None):
        """_summary_

        Args:
            id (int, optional): _description_. Defaults to None.
        """
        if (id==None):
            id = type(self).count()+1
        self.id=id
    
    def __str__(self):
        out=f"<Object from table '{type(self).get_table_name()}' : "
        for key,value in self.__dict__.items():
            out += f"{key} -> '{value}' | "
        return out[:-2]+">"
    
    def __eq__(self,other)->bool:
        if (type(self)!=type(other)):
            return False
        else:
            if (self.id != other.id):
                return False
            for key in type(self).get_property_names():
                if (other.__getattribute__(key)!=self.__getattribute__(key)):
                    return False
            return True
        
    @classmethod
    def show(cls,objects:list=None)->None:
        """prints down the table

        Args:
            objects (list, optional): list of class objects that should be printed. Defaults to None (cls.all() is printed if so).
        """
        if objects==None:
            objects=cls.all()
        print(f"TABLE {cls.get_table_name()} :")
        for cls_obj in objects:
            print("    ",end="")
            print(cls_obj)
        print("END")
    
    @saveDB
    def delete(self):
        """deletes the line corresponding to id in the database (by setting None in each of it's cells)
        """
        ws = type(self).get_table()
        for col,value in enumerate(self.__dict__.values(),start=1):
            ws.cell(self.id,col).value=None
    
    @saveDB
    def save(self):
        """saves the object in the line corresponding to the id (overwrites what was precedently on it)
        """
        ws = type(self).get_table()
        for col,value in enumerate(self.__dict__.values(),start=1):
            ws.cell(self.id,col).value=value
    
    @classmethod
    def create(cls):
        print(f"Creating table '{cls.get_table_name()}'...")
        if not Database.table_exists(cls.get_table_name()):
            Database.add_table(cls.get_table_name())
    
    @classmethod
    def drop(cls):
        Database.drop_table(cls.get_table_name())
    
    @classmethod
    def delete(cls):
        Database.delete_table(cls.get_table_name())
    
    @classmethod
    def count(cls)->int:
        """
        Returns:
            int: number of lines of the sheet (next empty line is count+1)
        """
        out = cls.get_table().max_row
        if (out==1 and cls.get_table().cell(1,1).value==None):
            out=0
        return out   
    
    @classmethod
    def get_table(cls)->Ws:
        return Database.get_table(cls.get_table_name())
    
    @classmethod
    def get_table_name(cls)->str:
        return cls.__name__.lower()+"s"    
            
    
    @classmethod
    def get(cls:type,id:int):
        return cls.parse_row(id)
    
    @classmethod
    def all(cls:type)->list:
        out=[]
        for i in range(1,cls.count()+1):
            try:
                out.append(cls.get(i))
            except InvalidID:
                pass
        return out
    
    @classmethod
    def filter(cls:type,filter:Filter,objects:list=None)->list:
        """

        Args:
            objects (list, optional): list of class objects (or id's, invalid id's will be ignored) you wan't to filter. Defaults to None for all objects of the table.
            filter (Filter, optional): Defaults to Filter().

        Returns:
            list: list of objects that correspond to the filter
        """
        if objects==None:
            objects=cls.all()
            
        out=[]
        for cls_obj in objects:
            if type(cls_obj)==int:
                try:
                    cls_obj = cls.get(cls_obj)
                except InvalidID:
                    continue
                
            if filter(cls_obj):
                out.append(cls_obj)
        return out
    
    @staticmethod
    def sort(objects:list,*args,function=None,desc:bool=False)->None:
        """in place modification !

        Args:
            objects (list): list of class objects (not id's) you wan't to sort
            function (function:cls_obj->int): list will be orderd by this function
            *args (string): list will then be orderd by the attributes you have set there

        Returns:
            list: objects in the right order
        """
        #list.sort sorts by lexicographic order !
        
        def func(cls_obj):
            out=[]
            for arg in args:
                out.append(cls_obj.__getattribute__(arg))
            if function!=None:
                out.append(function(cls_obj))
            return out
            
        objects.sort(key=func,reverse=desc)
        return
    
    @staticmethod
    def get_property_names()->list:
        raise IncompleteClassDefinition("get_property_names (@staticmethod)")
    
    @classmethod
    def parse_row(cls:type,id:int):
        ws = cls.get_table()
        if ws.cell(id,1).value==None:
            raise InvalidID()
        self = cls.__new__(cls)
        for col,property in enumerate(["id"]+cls.get_property_names(),start=1):
            self.__setattr__(property,ws.cell(id,col).value)
        return self

            
class Exemple(Model):
    
    test="test"
    
    def __init__(self,name:str):
        super().__init__()
        self.name=name
    
    @staticmethod
    def get_property_names() -> list:
        return ["name"]
    

if __name__=="__main__":
    
    restart=True
    
    if restart:
        Database.reset()
        Database.start()
        Exemple.create()
        
        for i in range(100):
            Exemple(f"Exemple n°{i}").save()
    
    Database.start()
    
    print(Exemple.get(20))
    
    
    #Exemple.show(Exemple.filter(filter=Filter(lambda x:x.id>=90)))
    
    Exemple.show(Exemple.filter(Filter(lambda x:x.id<3,id=1)))
    
    objects= Exemple.filter(Filter(lambda x:x.id>=90))
    Exemple.sort(objects,"id","name",desc=True)
    Exemple.show(objects)
    
    
    
    